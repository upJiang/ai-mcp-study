"""MCP 数据导出工具 - 支持 CSV/JSON/Excel 格式"""

import asyncio
import csv
import json
import logging
from datetime import datetime
from pathlib import Path
from typing import Any, Literal

from mcp.server import Server

logger = logging.getLogger(__name__)

ExportFormat = Literal["csv", "json", "excel"]


def register_export_tools(server: Server, db_ops: Any) -> None:
    """
    注册数据导出相关的 MCP Tools

    Args:
        server: MCP Server 实例
        db_ops: DatabaseOperations 实例
    """

    @server.tool()
    async def export_query_results(
        database: str,
        query: str,
        format: ExportFormat = "csv",
        filename: str | None = None,
        limit: int = 1000
    ) -> dict[str, Any]:
        """
        导出查询结果到文件

        Args:
            database: 数据库名称
            query: SQL 查询语句（仅支持 SELECT）
            format: 导出格式（csv/json/excel）
            filename: 文件名（可选，默认自动生成）
            limit: 最大导出行数（1-10000，默认 1000）

        Returns:
            导出结果信息（包含文件路径、行数、大小等）
        """
        try:
            # 参数验证
            if limit < 1 or limit > 10000:
                return {
                    "success": False,
                    "error": "limit 必须在 1-10000 之间"
                }

            if format not in ("csv", "json", "excel"):
                return {
                    "success": False,
                    "error": f"不支持的格式: {format}，支持的格式: csv, json, excel"
                }

            # 执行查询
            results = await asyncio.to_thread(
                db_ops.query_database,
                database,
                query,
                limit
            )

            if not results:
                return {
                    "success": False,
                    "error": "查询结果为空，无法导出"
                }

            # 生成文件名
            if not filename:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                filename = f"export_{database}_{timestamp}"

            # 移除文件扩展名（如果有）
            filename = Path(filename).stem

            # 导出目录（使用 tmp/ 目录，符合文件组织规则）
            export_dir = Path(__file__).parent.parent / "tmp" / "exports"
            export_dir.mkdir(parents=True, exist_ok=True)

            # 根据格式导出
            if format == "csv":
                file_path = await _export_to_csv(results, export_dir, filename)
            elif format == "json":
                file_path = await _export_to_json(results, export_dir, filename)
            elif format == "excel":
                file_path = await _export_to_excel(results, export_dir, filename)
            else:
                return {
                    "success": False,
                    "error": f"不支持的格式: {format}"
                }

            # 获取文件大小
            file_size = file_path.stat().st_size

            logger.info(
                f"导出成功: {database}.{query[:50]}... -> {file_path} "
                f"({len(results)} 行, {file_size} 字节)"
            )

            return {
                "success": True,
                "format": format,
                "file_path": str(file_path),
                "row_count": len(results),
                "file_size": file_size,
                "file_size_human": _format_size(file_size),
                "database": database,
                "query": query[:100] + ("..." if len(query) > 100 else "")
            }

        except Exception as e:
            logger.error(f"导出失败 [{database}]: {str(e)}")
            return {
                "success": False,
                "error": str(e),
                "database": database
            }


async def _export_to_csv(
    results: list[dict[str, Any]],
    export_dir: Path,
    filename: str
) -> Path:
    """导出为 CSV 格式"""
    file_path = export_dir / f"{filename}.csv"

    def _write_csv():
        with open(file_path, "w", newline="", encoding="utf-8-sig") as f:
            if results:
                writer = csv.DictWriter(f, fieldnames=results[0].keys())
                writer.writeheader()
                writer.writerows(results)

    await asyncio.to_thread(_write_csv)
    return file_path


async def _export_to_json(
    results: list[dict[str, Any]],
    export_dir: Path,
    filename: str
) -> Path:
    """导出为 JSON 格式"""
    file_path = export_dir / f"{filename}.json"

    def _write_json():
        with open(file_path, "w", encoding="utf-8") as f:
            json.dump(
                {
                    "export_time": datetime.now().isoformat(),
                    "row_count": len(results),
                    "data": results
                },
                f,
                ensure_ascii=False,
                indent=2,
                default=str  # 处理 datetime 等特殊类型
            )

    await asyncio.to_thread(_write_json)
    return file_path


async def _export_to_excel(
    results: list[dict[str, Any]],
    export_dir: Path,
    filename: str
) -> Path:
    """导出为 Excel 格式"""
    file_path = export_dir / f"{filename}.xlsx"

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        raise ImportError(
            "导出 Excel 需要安装 openpyxl: pip install openpyxl"
        )

    def _write_excel():
        # 创建工作簿
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Export Data"

        if not results:
            wb.save(file_path)
            return

        # 写入表头
        headers = list(results[0].keys())
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            # 表头样式
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

        # 写入数据
        for row_idx, row_data in enumerate(results, 2):
            for col_idx, header in enumerate(headers, 1):
                value = row_data.get(header)
                # 处理特殊类型
                if value is not None and not isinstance(value, (str, int, float, bool)):
                    value = str(value)
                ws.cell(row=row_idx, column=col_idx, value=value)

        # 自动调整列宽
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = min(max_length + 2, 50)  # 最大宽度 50
            ws.column_dimensions[column_letter].width = adjusted_width

        wb.save(file_path)

    await asyncio.to_thread(_write_excel)
    return file_path


def _format_size(size_bytes: int) -> str:
    """格式化文件大小为人类可读格式"""
    for unit in ["B", "KB", "MB", "GB"]:
        if size_bytes < 1024.0:
            return f"{size_bytes:.1f} {unit}"
        size_bytes /= 1024.0
    return f"{size_bytes:.1f} TB"
